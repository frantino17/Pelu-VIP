"""
Peluquería VIP - Aplicación de Simulación
Interfaz gráfica con PyQt5
"""

import sys
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QLabel, QSpinBox, 
                             QTableWidget, QTableWidgetItem, QGroupBox, 
                             QGridLayout, QHeaderView, QProgressBar, QTextEdit,
                             QSplitter, QTabWidget, QMessageBox, QFileDialog)
from PyQt5.QtCore import Qt, QTimer, pyqtSignal, QThread
from PyQt5.QtGui import QFont, QColor
import random

from simulacion import SimulacionPeluqueria, EstadoPeluquero

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class SimulacionThread(QThread):
    """Thread para ejecutar la simulación sin bloquear la UI"""
    progreso = pyqtSignal(int, int)  # día actual, total días
    completado = pyqtSignal(dict)  # resultados
    dia_completado = pyqtSignal(dict, int)  # stats de un día, número de día
    
    def __init__(self, num_dias, tiempo_max, max_iteraciones, params_modelo):
        super().__init__()
        self.num_dias = num_dias
        self.tiempo_max = tiempo_max
        self.max_iteraciones = max_iteraciones
        self.params_modelo = params_modelo
        self.simulacion = SimulacionPeluqueria(**params_modelo)
    
    def run(self):
        resultados = []
        
        for dia in range(self.num_dias):
            stats = self.simulacion.simular_dia(self.tiempo_max, self.max_iteraciones)
            stats['dia'] = dia + 1
            resultados.append(stats)
            
            # Emitir evento de día completado (solo guardamos el último día)
            if dia == self.num_dias - 1:
                self.dia_completado.emit(stats, dia + 1)
            
            self.progreso.emit(dia + 1, self.num_dias)
        
        stats_agregadas = self.simulacion._calcular_estadisticas_agregadas(resultados)
        self.completado.emit(stats_agregadas)


class PeluqueriaVIPApp(QMainWindow):
    """Ventana principal de la aplicación"""
    
    def __init__(self):
        super().__init__()
        self.simulacion = None  # Se creará con parámetros al ejecutar
        self.resultados = None
        self.ultima_simulacion = None  # Guardar referencia a la última simulación
        self.init_ui()
    
    def init_ui(self):
        """Inicializa la interfaz de usuario"""
        self.setWindowTitle('Simulación Peluquería VIP')
        self.setGeometry(50, 50, 1600, 950)  # Ventana más grande
        
        # Maximizar ventana al inicio (con botones de control visibles)
        self.showMaximized()
        
        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Layout principal
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        
        # Título
        titulo = QLabel('🪒 PELUQUERÍA VIP - SIMULACIÓN 🪒')
        titulo.setAlignment(Qt.AlignCenter)
        font = QFont()
        font.setPointSize(24)  # Tamaño de fuente más grande
        font.setBold(True)
        titulo.setFont(font)
        main_layout.addWidget(titulo)
        
        # Panel de control
        control_panel = self._crear_panel_control()
        main_layout.addWidget(control_panel)
        
        # Barra de progreso
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)
        
        # Tabs para diferentes vistas
        tabs = QTabWidget()
        
        # Tab 1: Vector de Estado
        self.tab_vector = self._crear_tab_vector_estado()
        tabs.addTab(self.tab_vector, "📋 Vector de Estado")
        
        # Tab 2: Resultados agregados
        self.tab_resultados = self._crear_tab_resultados()
        tabs.addTab(self.tab_resultados, "📊 Resultados Agregados")
        
        # Tab 3: Resultados por día
        self.tab_diarios = self._crear_tab_diarios()
        tabs.addTab(self.tab_diarios, "📅 Resultados Diarios")
        
        # Tab 4: Información del modelo
        self.tab_info = self._crear_tab_informacion()
        tabs.addTab(self.tab_info, "ℹ️ Información del Modelo")
        
        main_layout.addWidget(tabs)
        
        # Actualizar información del modelo con valores iniciales
        self._actualizar_info_modelo()
        
        # Estilo
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f0f0f0;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #cccccc;
                border-radius: 6px;
                margin-top: 6px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px 18px;  
                font-size: 15px;  
                border-radius: 4px;
                min-width: 120px;  
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
            QTableWidget {
                gridline-color: #d0d0d0;
                background-color: white;
            }
        """)
    
    def _crear_panel_control(self):
        """Crea el panel de control"""
        group = QGroupBox("⚙️ Configuración de Simulación")
        
        # Aumentar tamaño de fuente del título del grupo
        font_titulo_grupo = QFont()
        font_titulo_grupo.setPointSize(14)
        font_titulo_grupo.setBold(True)
        group.setFont(font_titulo_grupo)
        
        layout = QGridLayout()
        
        # Fuente para los SpinBox
        font_spin = QFont()
        font_spin.setPointSize(12)
        
        # Fuente para los labels
        font_label = QFont()
        font_label.setPointSize(13)
        
        row = 0
        
        # ===== SECCIÓN: PARÁMETROS DE SIMULACIÓN =====
        lbl_sim = QLabel("<b>PARÁMETROS DE SIMULACIÓN:</b>")
        font_section = QFont()
        font_section.setPointSize(13)
        font_section.setBold(True)
        lbl_sim.setFont(font_section)
        layout.addWidget(lbl_sim, row, 0, 1, 4)
        row += 1
        
        # Número de días
        lbl_dias = QLabel("Número de días a simular:")
        lbl_dias.setFont(font_label)
        layout.addWidget(lbl_dias, row, 0)
        self.spin_dias = QSpinBox()
        self.spin_dias.setMinimum(1)
        self.spin_dias.setMaximum(10000)
        self.spin_dias.setValue(30)
        self.spin_dias.setSuffix(" días")
        self.spin_dias.setMinimumHeight(28)
        self.spin_dias.setFont(font_spin)
        layout.addWidget(self.spin_dias, row, 1)
        
        # Tiempo máximo por día
        lbl_tiempo = QLabel("Tiempo máximo por día (min):")
        lbl_tiempo.setFont(font_label)
        layout.addWidget(lbl_tiempo, row, 2)
        self.spin_tiempo_max = QSpinBox()
        self.spin_tiempo_max.setMinimum(480)
        self.spin_tiempo_max.setMaximum(2000)
        self.spin_tiempo_max.setValue(1000)
        self.spin_tiempo_max.setSuffix(" min")
        self.spin_tiempo_max.setMinimumHeight(28)
        self.spin_tiempo_max.setFont(font_spin)
        layout.addWidget(self.spin_tiempo_max, row, 3)
        row += 1
        
        # Máximo de iteraciones
        lbl_iter = QLabel("Máximo de iteraciones:")
        lbl_iter.setFont(font_label)
        layout.addWidget(lbl_iter, row, 0)
        self.spin_max_iter = QSpinBox()
        self.spin_max_iter.setMinimum(100)
        self.spin_max_iter.setMaximum(100000)
        self.spin_max_iter.setValue(10000)
        self.spin_max_iter.setSingleStep(1000)
        self.spin_max_iter.setMinimumHeight(28)
        self.spin_max_iter.setFont(font_spin)
        layout.addWidget(self.spin_max_iter, row, 1, 1, 3)
        row += 1
        
        # ===== SECCIÓN: PARÁMETROS DEL MODELO =====
        lbl_modelo = QLabel("<b>PARÁMETROS DEL MODELO (valores en ROJO del enunciado):</b>")
        lbl_modelo.setFont(font_section)
        layout.addWidget(lbl_modelo, row, 0, 1, 4)
        row += 1
        
        # Subsección: Aprendiz
        lbl_apr = QLabel("<span style='color:red;'>Aprendiz:</span>")
        lbl_apr.setFont(font_label)
        layout.addWidget(lbl_apr, row, 0)
        row += 1
        
        lbl1 = QLabel("  Probabilidad (%):")
        lbl1.setFont(font_label)
        layout.addWidget(lbl1, row, 0)
        self.spin_prob_aprendiz = QSpinBox()
        self.spin_prob_aprendiz.setMinimum(0)
        self.spin_prob_aprendiz.setMaximum(100)
        self.spin_prob_aprendiz.setValue(15)
        self.spin_prob_aprendiz.setSuffix(" %")
        self.spin_prob_aprendiz.setFont(font_spin)
        self.spin_prob_aprendiz.valueChanged.connect(self._actualizar_prob_vet_b)
        layout.addWidget(self.spin_prob_aprendiz, row, 1)
        
        lbl2 = QLabel("  Tiempo U(min, max):")
        lbl2.setFont(font_label)
        layout.addWidget(lbl2, row, 2)
        self.spin_tiempo_min_apr = QSpinBox()
        self.spin_tiempo_min_apr.setMinimum(1)
        self.spin_tiempo_min_apr.setMaximum(100)
        self.spin_tiempo_min_apr.setValue(20)
        self.spin_tiempo_min_apr.setFont(font_spin)
        layout.addWidget(self.spin_tiempo_min_apr, row, 3)
        self.spin_tiempo_max_apr = QSpinBox()
        self.spin_tiempo_max_apr.setMinimum(1)
        self.spin_tiempo_max_apr.setMaximum(100)
        self.spin_tiempo_max_apr.setValue(30)
        self.spin_tiempo_max_apr.setFont(font_spin)
        layout.addWidget(self.spin_tiempo_max_apr, row, 4)
        row += 1
        
        # Subsección: Veterano A
        lbl_veta = QLabel("<span style='color:red;'>Veterano A:</span>")
        lbl_veta.setFont(font_label)
        layout.addWidget(lbl_veta, row, 0)
        row += 1
        
        lbl4 = QLabel("  Probabilidad (%):")
        lbl4.setFont(font_label)
        layout.addWidget(lbl4, row, 0)
        self.spin_prob_vet_a = QSpinBox()
        self.spin_prob_vet_a.setMinimum(0)
        self.spin_prob_vet_a.setMaximum(100)
        self.spin_prob_vet_a.setValue(45)
        self.spin_prob_vet_a.setSuffix(" %")
        self.spin_prob_vet_a.setFont(font_spin)
        self.spin_prob_vet_a.valueChanged.connect(self._actualizar_prob_vet_b)
        layout.addWidget(self.spin_prob_vet_a, row, 1)
        
        lbl5 = QLabel("  Tiempo U(min, max):")
        lbl5.setFont(font_label)
        layout.addWidget(lbl5, row, 2)
        self.spin_tiempo_min_vet_a = QSpinBox()
        self.spin_tiempo_min_vet_a.setMinimum(1)
        self.spin_tiempo_min_vet_a.setMaximum(100)
        self.spin_tiempo_min_vet_a.setValue(11)
        self.spin_tiempo_min_vet_a.setFont(font_spin)
        layout.addWidget(self.spin_tiempo_min_vet_a, row, 3)
        self.spin_tiempo_max_vet_a = QSpinBox()
        self.spin_tiempo_max_vet_a.setMinimum(1)
        self.spin_tiempo_max_vet_a.setMaximum(100)
        self.spin_tiempo_max_vet_a.setValue(13)
        self.spin_tiempo_max_vet_a.setFont(font_spin)
        layout.addWidget(self.spin_tiempo_max_vet_a, row, 4)
        row += 1
        
        # Subsección: Veterano B
        lbl_vetb = QLabel("<span style='color:red;'>Veterano B:</span>")
        lbl_vetb.setFont(font_label)
        layout.addWidget(lbl_vetb, row, 0)
        row += 1
        
        lbl7 = QLabel("  Probabilidad (%):")
        lbl7.setFont(font_label)
        layout.addWidget(lbl7, row, 0)
        # Label para mostrar la probabilidad calculada (no editable)
        self.lbl_prob_vet_b = QLabel("<span style='color:green; font-weight:bold;'>40 %</span> (Calculada automáticamente)")
        self.lbl_prob_vet_b.setFont(font_label)
        layout.addWidget(self.lbl_prob_vet_b, row, 1, 1, 4)
        row += 1
        
        lbl8 = QLabel("  Tiempo U(min, max):")
        lbl8.setFont(font_label)
        layout.addWidget(lbl8, row, 0, 1, 2)
        self.spin_tiempo_min_vet_b = QSpinBox()
        self.spin_tiempo_min_vet_b.setMinimum(1)
        self.spin_tiempo_min_vet_b.setMaximum(100)
        self.spin_tiempo_min_vet_b.setValue(12)
        self.spin_tiempo_min_vet_b.setFont(font_spin)
        layout.addWidget(self.spin_tiempo_min_vet_b, row, 2)
        self.spin_tiempo_max_vet_b = QSpinBox()
        self.spin_tiempo_max_vet_b.setMinimum(1)
        self.spin_tiempo_max_vet_b.setMaximum(100)
        self.spin_tiempo_max_vet_b.setValue(18)
        self.spin_tiempo_max_vet_b.setFont(font_spin)
        layout.addWidget(self.spin_tiempo_max_vet_b, row, 3)
        row += 1
        
        # Subsección: Llegadas de clientes
        lbl_lleg = QLabel("<span style='color:red;'>Llegadas de Clientes:</span>")
        lbl_lleg.setFont(font_label)
        layout.addWidget(lbl_lleg, row, 0)
        row += 1
        
        lbl10 = QLabel("  Tiempo entre llegadas U(min, max):")
        lbl10.setFont(font_label)
        layout.addWidget(lbl10, row, 0, 1, 2)
        self.spin_llegada_min = QSpinBox()
        self.spin_llegada_min.setMinimum(1)
        self.spin_llegada_min.setMaximum(100)
        self.spin_llegada_min.setValue(2)
        self.spin_llegada_min.setSuffix(" min")
        self.spin_llegada_min.setFont(font_spin)
        layout.addWidget(self.spin_llegada_min, row, 2)
        self.spin_llegada_max = QSpinBox()
        self.spin_llegada_max.setMinimum(1)
        self.spin_llegada_max.setMaximum(100)
        self.spin_llegada_max.setValue(12)
        self.spin_llegada_max.setSuffix(" min")
        self.spin_llegada_max.setFont(font_spin)
        layout.addWidget(self.spin_llegada_max, row, 3)
        row += 1
        
        # Subsección: Tiempo para refrigerios (el único parámetro configurable)
        lbl_refrig = QLabel("<span style='color:red;'>Refrigerios:</span>")
        lbl_refrig.setFont(font_label)
        layout.addWidget(lbl_refrig, row, 0)
        row += 1
        
        lbl12 = QLabel("  Tiempo para refrigerio (min):")
        lbl12.setFont(font_label)
        layout.addWidget(lbl12, row, 0, 1, 2)
        self.spin_tiempo_refrig = QSpinBox()
        self.spin_tiempo_refrig.setMinimum(1)
        self.spin_tiempo_refrig.setMaximum(120)
        self.spin_tiempo_refrig.setValue(30)
        self.spin_tiempo_refrig.setSuffix(" min")
        self.spin_tiempo_refrig.setFont(font_spin)
        layout.addWidget(self.spin_tiempo_refrig, row, 2)
        row += 1
        
        # Separador
        lbl_sep = QLabel("<b>Filtros Vector de Estado:</b>")
        font_sep = QFont()
        font_sep.setPointSize(13)
        font_sep.setBold(True)
        lbl_sep.setFont(font_sep)
        layout.addWidget(lbl_sep, row, 0, 1, 4)
        row += 1
        
        # Hora inicio (j)
        lbl14 = QLabel("Desde hora (j) en minutos:")
        lbl14.setFont(font_label)
        layout.addWidget(lbl14, row, 0)
        self.spin_hora_inicio = QSpinBox()
        self.spin_hora_inicio.setMinimum(0)
        self.spin_hora_inicio.setMaximum(2000)
        self.spin_hora_inicio.setValue(0)
        self.spin_hora_inicio.setSuffix(" min")
        self.spin_hora_inicio.setMinimumHeight(28)
        self.spin_hora_inicio.setFont(font_spin)
        layout.addWidget(self.spin_hora_inicio, row, 1)
        
        # Cantidad de filas (i)
        lbl15 = QLabel("Mostrar (i) filas:")
        lbl15.setFont(font_label)
        layout.addWidget(lbl15, row, 2)
        self.spin_num_filas = QSpinBox()
        self.spin_num_filas.setMinimum(1)
        self.spin_num_filas.setMaximum(10000)
        self.spin_num_filas.setValue(50)
        self.spin_num_filas.setSuffix(" filas")
        self.spin_num_filas.setMinimumHeight(28)
        self.spin_num_filas.setFont(font_spin)
        layout.addWidget(self.spin_num_filas, row, 3)
        row += 1
        
        # Botones
        btn_layout = QHBoxLayout()
        
        self.btn_simular = QPushButton("▶️ Ejecutar Simulación")
        self.btn_simular.clicked.connect(self.ejecutar_simulacion)
        btn_layout.addWidget(self.btn_simular)
        
        self.btn_actualizar_vector = QPushButton("🔄 Actualizar Vector")
        self.btn_actualizar_vector.clicked.connect(self.actualizar_vector_estado)
        self.btn_actualizar_vector.setEnabled(False)
        btn_layout.addWidget(self.btn_actualizar_vector)
        
        self.btn_exportar = QPushButton("📊 Exportar a Excel")
        self.btn_exportar.clicked.connect(self.exportar_simulacion)
        self.btn_exportar.setEnabled(False)
        btn_layout.addWidget(self.btn_exportar)
        
        self.btn_limpiar = QPushButton("🗑️ Limpiar Resultados")
        self.btn_limpiar.clicked.connect(self.limpiar_resultados)
        btn_layout.addWidget(self.btn_limpiar)
        
        layout.addLayout(btn_layout, row, 0, 1, 5)
        
        group.setLayout(layout)
        return group
    
    def _crear_tab_vector_estado(self):
        """Crea el tab del vector de estado"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Info label
        info_label = QLabel(
            "<span style='font-size:13px;'>Vector de Estado: Muestra <b>i</b> iteraciones desde la hora <b>j</b> + última fila siempre. "
            "Modifica los parámetros arriba y presiona 'Actualizar Vector'.</span>"
        )
        info_label.setWordWrap(True)
        info_label.setStyleSheet("background-color: #e3f2fd; padding: 12px; border-radius: 4px;")
        layout.addWidget(info_label)
        
        # Tabla del vector de estado
        self.tabla_vector = QTableWidget()
        self.tabla_vector.setAlternatingRowColors(True)
        
        # Configurar fuente mucho más grande para la tabla
        font_tabla = QFont()
        font_tabla.setPointSize(12)  # Aumentado de 10 a 12
        self.tabla_vector.setFont(font_tabla)
        
        # Configurar columnas
        columnas = [
            "Iter", "Reloj\n(min)", "Evento", 
            "RND\nLlegada", "RND\nAsig", "RND\nServicio",
            "Prox.\nLlegada", "Prox.\nFin Apr", "Prox.\nFin VetA", "Prox.\nFin VetB",
            "Estado\nAprendiz", "Cliente\nAprendiz", "Cola\nAprendiz",
            "Estado\nVet A", "Cliente\nVet A", "Cola\nVet A",
            "Estado\nVet B", "Cliente\nVet B", "Cola\nVet B",
            "Clientes\nAtendidos", "Recaud.\nAcum", "Costo\nRefrig", "Refrig.\nEntregados", "Max\nCola"
        ]
        
        self.tabla_vector.setColumnCount(len(columnas))
        self.tabla_vector.setHorizontalHeaderLabels(columnas)
        
        # Configurar fuente mucho más grande para el header
        header_font = QFont()
        header_font.setPointSize(11)  # Aumentado de 9 a 11
        header_font.setBold(True)
        self.tabla_vector.horizontalHeader().setFont(header_font)
        
        # Ajustar altura de las filas mucho más grande
        self.tabla_vector.verticalHeader().setDefaultSectionSize(35)  # Aumentado de 28 a 35
        
        # Ajustar tamaño de columnas
        header = self.tabla_vector.horizontalHeader()
        for i in range(len(columnas)):
            if i < 3:  # Primeras columnas más anchas
                header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
            else:
                header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        
        layout.addWidget(self.tabla_vector)
        widget.setLayout(layout)
        return widget
    
    def _crear_tab_resultados(self):
        """Crea el tab de resultados agregados"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Estadísticas principales
        stats_group = QGroupBox("📈 Estadísticas Generales")
        
        # Aumentar tamaño de fuente del título del grupo
        font_titulo_stats = QFont()
        font_titulo_stats.setPointSize(14)
        font_titulo_stats.setBold(True)
        stats_group.setFont(font_titulo_stats)
        
        stats_layout = QGridLayout()
        
        # Fuente para labels de las estadísticas
        font_label_stat = QFont()
        font_label_stat.setPointSize(13)
        
        # Labels para resultados
        self.lbl_num_dias = self._crear_label_resultado("N/A")
        self.lbl_recaudacion_prom = self._crear_label_resultado("N/A")
        self.lbl_recaudacion_min = self._crear_label_resultado("N/A")
        self.lbl_recaudacion_max = self._crear_label_resultado("N/A")
        self.lbl_ganancia_prom = self._crear_label_resultado("N/A")
        
        lbl_stat1 = QLabel("Días simulados:")
        lbl_stat1.setFont(font_label_stat)
        stats_layout.addWidget(lbl_stat1, 0, 0)
        stats_layout.addWidget(self.lbl_num_dias, 0, 1)
        
        lbl_stat2 = QLabel("Recaudación promedio diaria:")
        lbl_stat2.setFont(font_label_stat)
        stats_layout.addWidget(lbl_stat2, 1, 0)
        stats_layout.addWidget(self.lbl_recaudacion_prom, 1, 1)
        
        lbl_stat3 = QLabel("Recaudación mínima:")
        lbl_stat3.setFont(font_label_stat)
        stats_layout.addWidget(lbl_stat3, 2, 0)
        stats_layout.addWidget(self.lbl_recaudacion_min, 2, 1)
        
        lbl_stat4 = QLabel("Recaudación máxima:")
        lbl_stat4.setFont(font_label_stat)
        stats_layout.addWidget(lbl_stat4, 3, 0)
        stats_layout.addWidget(self.lbl_recaudacion_max, 3, 1)
        
        lbl_stat5 = QLabel("Ganancia neta promedio:")
        lbl_stat5.setFont(font_label_stat)
        stats_layout.addWidget(lbl_stat5, 4, 0)
        stats_layout.addWidget(self.lbl_ganancia_prom, 4, 1)
        
        stats_group.setLayout(stats_layout)
        layout.addWidget(stats_group)
        
        # Respuestas a las preguntas
        font_titulo_resp = QFont()
        font_titulo_resp.setPointSize(14)
        font_titulo_resp.setBold(True)
        
        respuestas_group = QGroupBox("❓ Respuestas del Modelo")
        respuestas_group.setFont(font_titulo_resp)
        respuestas_layout = QVBoxLayout()
        
        self.lbl_resp1 = self._crear_label_respuesta("¿Cuál es el promedio de recaudación diaria?", "N/A")
        self.lbl_resp2 = self._crear_label_respuesta("¿Cantidad de sillas necesarias?", "N/A")
        self.lbl_resp3 = self._crear_label_respuesta("¿Probabilidad de 5+ refrigerios en un día?", "N/A")
        
        respuestas_layout.addWidget(self.lbl_resp1)
        respuestas_layout.addWidget(self.lbl_resp2)
        respuestas_layout.addWidget(self.lbl_resp3)
        
        respuestas_group.setLayout(respuestas_layout)
        layout.addWidget(respuestas_group)
        
        # Estadísticas adicionales
        font_titulo_adic = QFont()
        font_titulo_adic.setPointSize(14)
        font_titulo_adic.setBold(True)
        
        font_label_adic = QFont()
        font_label_adic.setPointSize(13)
        
        adicional_group = QGroupBox("📊 Información Adicional")
        adicional_group.setFont(font_titulo_adic)
        adicional_layout = QGridLayout()
        
        self.lbl_sillas_prom = self._crear_label_resultado("N/A")
        self.lbl_refrig_prom = self._crear_label_resultado("N/A")
        self.lbl_dias_5_mas = self._crear_label_resultado("N/A")
        
        lbl_adic1 = QLabel("Sillas promedio necesarias:")
        lbl_adic1.setFont(font_label_adic)
        adicional_layout.addWidget(lbl_adic1, 0, 0)
        adicional_layout.addWidget(self.lbl_sillas_prom, 0, 1)
        
        lbl_adic2 = QLabel("Refrigerios promedio por día:")
        lbl_adic2.setFont(font_label_adic)
        adicional_layout.addWidget(lbl_adic2, 1, 0)
        adicional_layout.addWidget(self.lbl_refrig_prom, 1, 1)
        
        lbl_adic3 = QLabel("Días con 5 o más refrigerios:")
        lbl_adic3.setFont(font_label_adic)
        adicional_layout.addWidget(lbl_adic3, 2, 0)
        adicional_layout.addWidget(self.lbl_dias_5_mas, 2, 1)
        
        adicional_group.setLayout(adicional_layout)
        layout.addWidget(adicional_group)
        
        layout.addStretch()
        widget.setLayout(layout)
        return widget
    
    def _crear_tab_diarios(self):
        """Crea el tab de resultados diarios"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Tabla de resultados diarios
        self.tabla_diarios = QTableWidget()
        self.tabla_diarios.setColumnCount(7)
        self.tabla_diarios.setHorizontalHeaderLabels([
            'Día', 'Recaudación', 'Costo Refrig.', 'Ganancia Neta',
            'Clientes', 'Refrigerios', 'Sillas Usadas'
        ])
        
        # Configurar fuente mucho más grande
        font_diarios = QFont()
        font_diarios.setPointSize(13)  # Aumentado de 11 a 13
        self.tabla_diarios.setFont(font_diarios)
        
        header = self.tabla_diarios.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        
        # Fuente para header
        header_font = QFont()
        header_font.setPointSize(12)  # Aumentado de 10 a 12
        header_font.setBold(True)
        header.setFont(header_font)
        
        self.tabla_diarios.setAlternatingRowColors(True)
        self.tabla_diarios.verticalHeader().setDefaultSectionSize(35)  # Aumentado de 30 a 35
        
        layout.addWidget(self.tabla_diarios)
        widget.setLayout(layout)
        return widget
    
    def _crear_tab_informacion(self):
        """Crea el tab de información"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        self.info_text = QTextEdit()
        self.info_text.setReadOnly(True)
        self._actualizar_info_modelo()
        
        layout.addWidget(self.info_text)
        widget.setLayout(layout)
        return widget
    
    def _actualizar_info_modelo(self):
        """Actualiza la información del modelo con los valores actuales"""
        # Obtener valores actuales de los controles parametrizables (si existen)
        prob_apr = self.spin_prob_aprendiz.value() if hasattr(self, 'spin_prob_aprendiz') else 15
        t_min_apr = self.spin_tiempo_min_apr.value() if hasattr(self, 'spin_tiempo_min_apr') else 20
        t_max_apr = self.spin_tiempo_max_apr.value() if hasattr(self, 'spin_tiempo_max_apr') else 30
        
        prob_vet_a = self.spin_prob_vet_a.value() if hasattr(self, 'spin_prob_vet_a') else 45
        t_min_vet_a = self.spin_tiempo_min_vet_a.value() if hasattr(self, 'spin_tiempo_min_vet_a') else 11
        t_max_vet_a = self.spin_tiempo_max_vet_a.value() if hasattr(self, 'spin_tiempo_max_vet_a') else 13
        
        # Calcular probabilidad del Veterano B (depende de las otras dos)
        prob_vet_b = 100 - prob_apr - prob_vet_a
        t_min_vet_b = self.spin_tiempo_min_vet_b.value() if hasattr(self, 'spin_tiempo_min_vet_b') else 12
        t_max_vet_b = self.spin_tiempo_max_vet_b.value() if hasattr(self, 'spin_tiempo_max_vet_b') else 18
        
        lleg_min = self.spin_llegada_min.value() if hasattr(self, 'spin_llegada_min') else 2
        lleg_max = self.spin_llegada_max.value() if hasattr(self, 'spin_llegada_max') else 12
        
        t_refrig = self.spin_tiempo_refrig.value() if hasattr(self, 'spin_tiempo_refrig') else 30
        
        # Valores CONSTANTES (NO parametrizables - valores fijos del enunciado)
        tarifa_apr = 18000
        tarifa_vet_a = 32500
        tarifa_vet_b = 32500
        jornada = 8
        costo_refrig = 5500
        
        html_content = f"""
        <h2>📋 Modelo de Simulación - Peluquería VIP</h2>
        
        <h3>🧑‍💼 Peluqueros:</h3>
        <ul>
            <li><b>Aprendiz:</b> Atiende <span style='color:red;font-weight:bold'>{prob_apr}%</span> de clientes, 
                U(<span style='color:red;font-weight:bold'>{t_min_apr}, {t_max_apr}</span>) min, 
                $<span style='font-weight:bold'>{tarifa_apr:,}</span> por corte <span style='color:green;'>(FIJO)</span></li>
            <li><b>Veterano A:</b> Atiende <span style='color:red;font-weight:bold'>{prob_vet_a}%</span> de clientes, 
                U(<span style='color:red;font-weight:bold'>{t_min_vet_a}, {t_max_vet_a}</span>) min, 
                $<span style='font-weight:bold'>{tarifa_vet_a:,}</span> por corte <span style='color:green;'>(FIJO)</span></li>
            <li><b>Veterano B:</b> Atiende <span style='color:green;font-weight:bold'>{prob_vet_b}%</span> de clientes <span style='color:green;'>(CALCULADO = 100% - {prob_apr}% - {prob_vet_a}%)</span>, 
                U(<span style='color:red;font-weight:bold'>{t_min_vet_b}, {t_max_vet_b}</span>) min, 
                $<span style='font-weight:bold'>{tarifa_vet_b:,}</span> por corte <span style='color:green;'>(FIJO)</span></li>
        </ul>
        
        <h3>👥 Llegada de Clientes:</h3>
        <ul>
            <li>Tiempo entre llegadas: U(<span style='color:red;font-weight:bold'>{lleg_min}, {lleg_max}</span>) minutos</li>
            <li>Se receptan durante <span style='font-weight:bold'>{jornada} horas</span> (<span style='font-weight:bold'>{jornada*60}</span> minutos) <span style='color:green;'>(FIJO)</span></li>
            <li>Se trabaja hasta que no queden clientes</li>
        </ul>
        
        <h3>🥤 Política de Refrigerios:</h3>
        <ul>
            <li>Si un cliente espera más de <span style='color:red;font-weight:bold'>{t_refrig} minutos</span>, recibe un refrigerio</li>
            <li>Costo del refrigerio: $<span style='font-weight:bold'>{costo_refrig:,}</span> <span style='color:green;'>(FIJO)</span></li>
            <li>El cliente puede seguir esperando con su bebida</li>
        </ul>
        
        <h3>📐 Fórmulas y Cálculos:</h3>
        <h4>1. Recaudación Promedio Diaria:</h4>
        <p><b>Fórmula:</b> Recaudación_Promedio = Σ(Recaudación_i) / N</p>
        <ul>
            <li>Donde N = número de días simulados</li>
            <li>Recaudación_i = recaudación del día i</li>
            <li>Recaudación_i = Σ(Tarifa_peluquero por cada cliente atendido)</li>
        </ul>
        
        <h4>2. Cantidad de Sillas Necesarias:</h4>
        <p><b>Fórmula:</b> Sillas = MAX(Cola_Total_t) para todo t</p>
        <ul>
            <li>Cola_Total_t = número de clientes esperando en el tiempo t</li>
            <li>Se toma el máximo de todos los días simulados</li>
        </ul>
        
        <h4>3. Probabilidad de 5+ Refrigerios:</h4>
        <p><b>Fórmula:</b> P(Refrigerios ≥ 5) = Días_con_5_o_más / N</p>
        <ul>
            <li>Días_con_5_o_más = cantidad de días donde se entregaron 5 o más refrigerios</li>
            <li>N = total de días simulados</li>
        </ul>
        
        <h4>4. Ganancia Neta:</h4>
        <p><b>Fórmula:</b> Ganancia = Recaudación - Costo_Refrigerios</p>
        <ul>
            <li>Recaudación = Σ(Tarifa × Clientes_Atendidos)</li>
            <li>Costo_Refrigerios = N_Refrigerios × ${costo_refrig:,}</li>
        </ul>
        
        <h3>❓ Preguntas a Responder:</h3>
        <ol>
            <li>¿Cuál es el promedio de recaudación diaria de la peluquería?</li>
            <li>¿Qué cantidad de sillas son necesarias para que en ningún momento se encuentre un cliente de pie?</li>
            <li>¿Cuál es la probabilidad de que un día se le entregue refrigerio a 5 o más personas?</li>
        </ol>
        
        <h3>🎲 Método de Simulación:</h3>
        <p>Se utiliza simulación por <b>eventos discretos</b> donde se modelan:</p>
        <ul>
            <li><b>Llegadas de clientes:</b> Usando RND para tiempo entre llegadas U({lleg_min},{lleg_max})</li>
            <li><b>Asignación a peluqueros:</b> Usando RND con probabilidades acumuladas</li>
            <li><b>Tiempo de servicio:</b> Usando RND para U(min, max) según peluquero</li>
            <li><b>Entrega de refrigerios:</b> Cuando tiempo_espera > {t_refrig} minutos</li>
        </ul>
        
        <h3>📊 Vector de Estado:</h3>
        <p>El vector de estado muestra para cada iteración:</p>
        <ul>
            <li><b>Reloj:</b> Tiempo actual de simulación (minutos)</li>
            <li><b>Evento:</b> Tipo de evento procesado</li>
            <li><b>RNDs:</b> Números aleatorios usados en esta iteración</li>
            <li><b>Próximos Eventos:</b> Tiempos de próximos eventos programados</li>
            <li><b>Estado de Objetos:</b> Estado actual de cada peluquero y su cliente</li>
            <li><b>Colas:</b> Cantidad de clientes esperando por cada peluquero</li>
            <li><b>Acumuladores:</b> Variables agregadas (clientes atendidos, recaudación, etc.)</li>
        </ul>
        
        <p><i>Parámetros de Visualización:</i></p>
        <ul>
            <li><b>Tiempo máximo X:</b> Límite de tiempo de simulación por día</li>
            <li><b>Máximo iteraciones:</b> Máximo 100,000 iteraciones por día</li>
            <li><b>Hora inicio j:</b> Desde qué minuto mostrar el vector</li>
            <li><b>Cantidad i:</b> Cuántas filas mostrar desde j</li>
        </ul>
        
        <p><i>La última fila siempre se muestra (fondo amarillo) independiente de los filtros.</i></p>
        
        <p style='background-color:#fff3cd; padding:10px; border-radius:5px; margin-top:20px;'>
        <b>💡 Nota:</b> Los valores en <span style='color:red;font-weight:bold'>ROJO</span> son <b>parametrizables</b> 
        desde el panel de configuración superior. Los valores en <span style='color:green;font-weight:bold'>VERDE (FIJO)</span> 
        son <b>constantes del enunciado</b> y NO se pueden modificar.
        </p>
        """
        
        if hasattr(self, 'info_text'):
            self.info_text.setHtml(html_content)
    
    def _actualizar_prob_vet_b(self):
        """Actualiza la probabilidad del Veterano B basándose en las otras dos"""
        if hasattr(self, 'lbl_prob_vet_b'):
            prob_apr = self.spin_prob_aprendiz.value()
            prob_vet_a = self.spin_prob_vet_a.value()
            prob_vet_b = 100 - prob_apr - prob_vet_a
            
            # Validar que la probabilidad sea válida
            if prob_vet_b < 0:
                self.lbl_prob_vet_b.setText(
                    f"<span style='color:red; font-weight:bold;'>ERROR: {prob_vet_b} %</span> (Suma > 100%)"
                )
            else:
                self.lbl_prob_vet_b.setText(
                    f"<span style='color:green; font-weight:bold;'>{prob_vet_b} %</span> (Calculada = 100% - {prob_apr}% - {prob_vet_a}%)"
                )
    
    def _crear_label_resultado(self, texto_inicial):
        """Crea un label para mostrar resultados"""
        lbl = QLabel(texto_inicial)
        font = QFont()
        font.setPointSize(14)  # Aumentado de 12 a 14
        font.setBold(True)
        lbl.setFont(font)
        return lbl
    
    def _crear_label_respuesta(self, pregunta, respuesta):
        """Crea un label para pregunta-respuesta"""
        lbl = QLabel(f"<b style='font-size:15px;'>{pregunta}</b><br/><span style='font-size:20px; color:#2196F3;'>{respuesta}</span>")
        lbl.setWordWrap(True)
        return lbl
    
    def ejecutar_simulacion(self):
        """Ejecuta la simulación"""
        num_dias = self.spin_dias.value()
        tiempo_max = self.spin_tiempo_max.value()
        max_iter = self.spin_max_iter.value()
        
        # Recoger parámetros del modelo desde la UI
        params_modelo = {
            'prob_aprendiz': self.spin_prob_aprendiz.value() / 100.0,
            'tiempo_min_aprendiz': self.spin_tiempo_min_apr.value(),
            'tiempo_max_aprendiz': self.spin_tiempo_max_apr.value(),
            'prob_veterano_a': self.spin_prob_vet_a.value() / 100.0,
            'tiempo_min_vet_a': self.spin_tiempo_min_vet_a.value(),
            'tiempo_max_vet_a': self.spin_tiempo_max_vet_a.value(),
            'tiempo_min_vet_b': self.spin_tiempo_min_vet_b.value(),
            'tiempo_max_vet_b': self.spin_tiempo_max_vet_b.value(),
            'tiempo_llegada_min': self.spin_llegada_min.value(),
            'tiempo_llegada_max': self.spin_llegada_max.value(),
            'tiempo_refrigerio': self.spin_tiempo_refrig.value()
        }
        
        # Deshabilitar controles
        self.btn_simular.setEnabled(False)
        self.spin_dias.setEnabled(False)
        self.spin_tiempo_max.setEnabled(False)
        self.spin_max_iter.setEnabled(False)
        
        # Mostrar barra de progreso
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.progress_bar.setMaximum(num_dias)
        
        # Crear y ejecutar thread
        self.sim_thread = SimulacionThread(num_dias, tiempo_max, max_iter, params_modelo)
        self.sim_thread.progreso.connect(self._actualizar_progreso)
        self.sim_thread.completado.connect(self._mostrar_resultados)
        self.sim_thread.dia_completado.connect(self._guardar_ultima_simulacion)
        self.sim_thread.start()
    
    def _guardar_ultima_simulacion(self, stats, dia_num):
        """Guarda la referencia a la última simulación completa"""
        self.ultima_simulacion = self.sim_thread.simulacion
    
    def _actualizar_progreso(self, dia_actual, total_dias):
        """Actualiza la barra de progreso"""
        self.progress_bar.setValue(dia_actual)
        self.progress_bar.setFormat(f"Simulando día {dia_actual} de {total_dias}... ({dia_actual*100//total_dias}%)")
    
    def _mostrar_resultados(self, resultados):
        """Muestra los resultados de la simulación"""
        self.resultados = resultados
        
        # Actualizar estadísticas generales
        self.lbl_num_dias.setText(f"{resultados['num_dias']}")
        self.lbl_recaudacion_prom.setText(f"${resultados['recaudacion_promedio']:,.2f}")
        self.lbl_recaudacion_min.setText(f"${resultados['recaudacion_min']:,.2f}")
        self.lbl_recaudacion_max.setText(f"${resultados['recaudacion_max']:,.2f}")
        self.lbl_ganancia_prom.setText(f"${resultados['ganancia_promedio']:,.2f}")
        
        # Actualizar respuestas
        self.lbl_resp1.setText(
            f"<b style='font-size:15px;'>¿Cuál es el promedio de recaudación diaria?</b><br/>"
            f"<span style='font-size:20px; color:#2196F3; font-weight:bold;'>${resultados['recaudacion_promedio']:,.2f}</span>"
        )
        self.lbl_resp2.setText(
            f"<b style='font-size:15px;'>¿Cantidad de sillas necesarias?</b><br/>"
            f"<span style='font-size:20px; color:#2196F3; font-weight:bold;'>{resultados['max_sillas_necesarias']} sillas</span>"
        )
        self.lbl_resp3.setText(
            f"<b style='font-size:15px;'>¿Probabilidad de 5+ refrigerios en un día?</b><br/>"
            f"<span style='font-size:20px; color:#2196F3; font-weight:bold;'>{resultados['prob_5_o_mas_refrigerios']*100:.2f}%</span>"
        )
        
        # Actualizar estadísticas adicionales
        self.lbl_sillas_prom.setText(f"{resultados['sillas_promedio']:.2f}")
        self.lbl_refrig_prom.setText(f"{resultados['refrigerios_promedio']:.2f}")
        self.lbl_dias_5_mas.setText(
            f"{resultados['dias_5_o_mas_refrigerios']} de {resultados['num_dias']} "
            f"({resultados['dias_5_o_mas_refrigerios']*100/resultados['num_dias']:.1f}%)"
        )
        
        # Actualizar tabla de resultados diarios
        self._actualizar_tabla_diarios(resultados['resultados_diarios'])
        
        # Ocultar progreso y habilitar controles
        self.progress_bar.setVisible(False)
        self.btn_simular.setEnabled(True)
        self.spin_dias.setEnabled(True)
        self.spin_tiempo_max.setEnabled(True)
        self.spin_max_iter.setEnabled(True)
        self.btn_actualizar_vector.setEnabled(True)
        self.btn_exportar.setEnabled(True)
        
        # Actualizar el vector de estado
        self.actualizar_vector_estado()
        
        # Mostrar mensaje de completado
        QMessageBox.information(
            self,
            "Simulación Completada",
            f"Se han simulado {resultados['num_dias']} días exitosamente.\n\n"
            f"Recaudación promedio: ${resultados['recaudacion_promedio']:,.2f}\n"
            f"Sillas necesarias: {resultados['max_sillas_necesarias']}\n"
            f"Prob. 5+ refrigerios: {resultados['prob_5_o_mas_refrigerios']*100:.2f}%"
        )
    
    def _actualizar_tabla_diarios(self, resultados_diarios):
        """Actualiza la tabla de resultados diarios"""
        self.tabla_diarios.setRowCount(len(resultados_diarios))
        
        # Fuente para las celdas
        font_celda = QFont()
        font_celda.setPointSize(12)
        
        for i, resultado in enumerate(resultados_diarios):
            item0 = QTableWidgetItem(str(resultado['dia']))
            item0.setFont(font_celda)
            self.tabla_diarios.setItem(i, 0, item0)
            
            item1 = QTableWidgetItem(f"${resultado['recaudacion']:,.0f}")
            item1.setFont(font_celda)
            self.tabla_diarios.setItem(i, 1, item1)
            
            item2 = QTableWidgetItem(f"${resultado['costo_refrigerios']:,.0f}")
            item2.setFont(font_celda)
            self.tabla_diarios.setItem(i, 2, item2)
            
            item3 = QTableWidgetItem(f"${resultado['ganancia_neta']:,.0f}")
            item3.setFont(font_celda)
            self.tabla_diarios.setItem(i, 3, item3)
            
            item4 = QTableWidgetItem(str(resultado['clientes_atendidos']))
            item4.setFont(font_celda)
            self.tabla_diarios.setItem(i, 4, item4)
            
            item5 = QTableWidgetItem(str(resultado['clientes_con_refrigerio']))
            item5.setFont(font_celda)
            self.tabla_diarios.setItem(i, 5, item5)
            
            item6 = QTableWidgetItem(str(resultado['max_sillas_necesarias']))
            item6.setFont(font_celda)
            self.tabla_diarios.setItem(i, 6, item6)
            
            # Colorear filas con 5+ refrigerios
            if resultado['clientes_con_refrigerio'] >= 5:
                for j in range(7):
                    item = self.tabla_diarios.item(i, j)
                    if item:
                        item.setBackground(QColor(255, 235, 205))
    
    def actualizar_vector_estado(self):
        """Actualiza la visualización del vector de estado"""
        if not self.ultima_simulacion or not self.ultima_simulacion.vector_estado:
            return
        
        hora_inicio = self.spin_hora_inicio.value()
        num_filas = self.spin_num_filas.value()
        
        # Obtener filas filtradas
        filas = self.ultima_simulacion.obtener_vector_estado_filtrado(hora_inicio, num_filas)
        
        # Actualizar tabla
        self.tabla_vector.setRowCount(len(filas))
        
        # Fuente para las celdas - MÁS GRANDE
        font_celda = QFont()
        font_celda.setPointSize(11)  # Tamaño aumentado
        
        for i, fila in enumerate(filas):
            # Marcar última fila con fondo diferente
            es_ultima = (i == len(filas) - 1 and fila == self.ultima_simulacion.vector_estado[-1])
            
            # Columna 0: Iteración
            item = QTableWidgetItem(str(fila.iteracion))
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 0, item)
            
            # Columna 1: Reloj
            item = QTableWidgetItem(f"{fila.reloj:.2f}")
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 1, item)
            
            # Columna 2: Evento
            item = QTableWidgetItem(fila.evento)
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 2, item)
            
            # Columnas 3-5: RNDs
            rnd_llegada = fila.rnd_evento.get('llegada', '-')
            if rnd_llegada != '-':
                rnd_llegada = f"{rnd_llegada:.4f}"
            item = QTableWidgetItem(str(rnd_llegada))
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 3, item)
            
            rnd_asig = fila.rnd_evento.get('asignacion_peluquero', '-')
            if rnd_asig != '-':
                rnd_asig = f"{rnd_asig:.4f}"
            item = QTableWidgetItem(str(rnd_asig))
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 4, item)
            
            rnd_serv = fila.rnd_evento.get('tiempo_servicio', '-')
            if rnd_serv != '-':
                rnd_serv = f"{rnd_serv:.4f}"
            item = QTableWidgetItem(str(rnd_serv))
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 5, item)
            
            # Columnas 6-9: Próximos eventos
            item = QTableWidgetItem(f"{fila.proximo_llegada:.2f}" if fila.proximo_llegada > 0 else "-")
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 6, item)
            
            item = QTableWidgetItem(f"{fila.proximo_fin_aprendiz:.2f}" if fila.proximo_fin_aprendiz > 0 else "-")
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 7, item)
            
            item = QTableWidgetItem(f"{fila.proximo_fin_veterano_a:.2f}" if fila.proximo_fin_veterano_a > 0 else "-")
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 8, item)
            
            item = QTableWidgetItem(f"{fila.proximo_fin_veterano_b:.2f}" if fila.proximo_fin_veterano_b > 0 else "-")
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 9, item)
            
            # Columnas 10-12: Aprendiz
            item = QTableWidgetItem(fila.estado_aprendiz)
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 10, item)
            
            item = QTableWidgetItem(fila.cliente_aprendiz)
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 11, item)
            
            item = QTableWidgetItem(str(fila.cola_aprendiz))
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 12, item)
            
            # Columnas 13-15: Veterano A
            item = QTableWidgetItem(fila.estado_veterano_a)
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 13, item)
            
            item = QTableWidgetItem(fila.cliente_veterano_a)
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 14, item)
            
            item = QTableWidgetItem(str(fila.cola_veterano_a))
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 15, item)
            
            # Columnas 16-18: Veterano B
            item = QTableWidgetItem(fila.estado_veterano_b)
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 16, item)
            
            item = QTableWidgetItem(fila.cliente_veterano_b)
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 17, item)
            
            item = QTableWidgetItem(str(fila.cola_veterano_b))
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 18, item)
            
            # Columnas 19-23: Acumuladores
            item = QTableWidgetItem(str(fila.clientes_atendidos))
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 19, item)
            
            item = QTableWidgetItem(f"${fila.recaudacion_acum:,.0f}")
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 20, item)
            
            item = QTableWidgetItem(f"${fila.costo_refrigerios_acum:,.0f}")
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 21, item)
            
            item = QTableWidgetItem(str(fila.clientes_con_refrigerio))
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 22, item)
            
            item = QTableWidgetItem(str(fila.max_cola_total))
            item.setFont(font_celda)
            self.tabla_vector.setItem(i, 23, item)
            
            # Colorear última fila
            if es_ultima:
                for col in range(self.tabla_vector.columnCount()):
                    item = self.tabla_vector.item(i, col)
                    if item:
                        item.setBackground(QColor(255, 248, 220))  # Color amarillo claro
                        font = item.font()
                        font.setBold(True)
                        item.setFont(font)
    
    def limpiar_resultados(self):
        """Limpia los resultados mostrados"""
        self.resultados = None
        self.ultima_simulacion = None
        
        # Limpiar labels
        self.lbl_num_dias.setText("N/A")
        self.lbl_recaudacion_prom.setText("N/A")
        self.lbl_recaudacion_min.setText("N/A")
        self.lbl_recaudacion_max.setText("N/A")
        self.lbl_ganancia_prom.setText("N/A")
        self.lbl_sillas_prom.setText("N/A")
        self.lbl_refrig_prom.setText("N/A")
        self.lbl_dias_5_mas.setText("N/A")
        
        self.lbl_resp1.setText("<b style='font-size:15px;'>¿Cuál es el promedio de recaudación diaria?</b><br/><span style='font-size:20px; color:#2196F3;'>N/A</span>")
        self.lbl_resp2.setText("<b style='font-size:15px;'>¿Cantidad de sillas necesarias?</b><br/><span style='font-size:20px; color:#2196F3;'>N/A</span>")
        self.lbl_resp3.setText("<b style='font-size:15px;'>¿Probabilidad de 5+ refrigerios en un día?</b><br/><span style='font-size:20px; color:#2196F3;'>N/A</span>")
        
        # Limpiar tablas
        self.tabla_diarios.setRowCount(0)
        self.tabla_vector.setRowCount(0)
        
        # Deshabilitar botones
        self.btn_actualizar_vector.setEnabled(False)
        self.btn_exportar.setEnabled(False)
    
    def exportar_simulacion(self):
        """Exporta la simulación a un archivo Excel"""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(
                self,
                "Biblioteca No Disponible",
                "La biblioteca 'openpyxl' no está instalada.\n\n"
                "Para instalarla, ejecute:\n"
                "pip install openpyxl"
            )
            return
        
        if not self.resultados or not self.ultima_simulacion:
            QMessageBox.warning(
                self,
                "Sin Datos",
                "No hay datos de simulación para exportar.\n"
                "Por favor, ejecute una simulación primero."
            )
            return
        
        # Seleccionar archivo de destino
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_por_defecto = f"simulacion_peluqueria_{timestamp}.xlsx"
        
        archivo, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar Simulación",
            nombre_por_defecto,
            "Archivos Excel (*.xlsx);;Todos los archivos (*)"
        )
        
        if not archivo:
            return  # Usuario canceló
        
        try:
            self._exportar_a_excel(archivo)
            QMessageBox.information(
                self,
                "Exportación Exitosa",
                f"La simulación se ha exportado exitosamente a:\n{archivo}"
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error al Exportar",
                f"Ocurrió un error al exportar la simulación:\n{str(e)}"
            )
    
    def _exportar_a_excel(self, archivo):
        """Exporta los datos a un archivo Excel"""
        wb = Workbook()
        
        # Estilos
        titulo_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        titulo_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        center_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Hoja 1: Resumen General
        ws_resumen = wb.active
        ws_resumen.title = "Resumen General"
        
        # Título
        ws_resumen['A1'] = 'SIMULACIÓN PELUQUERÍA VIP - RESUMEN GENERAL'
        ws_resumen['A1'].font = titulo_font
        ws_resumen['A1'].fill = titulo_fill
        ws_resumen['A1'].alignment = center_alignment
        ws_resumen.merge_cells('A1:C1')
        
        # Fecha de exportación
        ws_resumen['A2'] = 'Fecha de Exportación:'
        ws_resumen['B2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Parámetros de simulación
        ws_resumen['A4'] = 'PARÁMETROS DE SIMULACIÓN'
        ws_resumen['A4'].font = Font(bold=True, size=12)
        ws_resumen['A5'] = 'Días simulados:'
        ws_resumen['B5'] = self.resultados['num_dias']
        ws_resumen['A6'] = 'Tiempo máximo por día (min):'
        ws_resumen['B6'] = self.spin_tiempo_max.value()
        ws_resumen['A7'] = 'Máximo iteraciones:'
        ws_resumen['B7'] = self.spin_max_iter.value()
        
        # Resultados agregados
        ws_resumen['A9'] = 'RESULTADOS AGREGADOS'
        ws_resumen['A9'].font = Font(bold=True, size=12)
        
        fila = 10
        ws_resumen[f'A{fila}'] = 'Métrica'
        ws_resumen[f'B{fila}'] = 'Valor'
        ws_resumen[f'A{fila}'].font = header_font
        ws_resumen[f'B{fila}'].font = header_font
        ws_resumen[f'A{fila}'].fill = header_fill
        ws_resumen[f'B{fila}'].fill = header_fill
        
        fila += 1
        ws_resumen[f'A{fila}'] = 'Recaudación Promedio Diaria'
        ws_resumen[f'B{fila}'] = self.resultados['recaudacion_promedio']
        ws_resumen[f'B{fila}'].number_format = '$#,##0.00'
        
        fila += 1
        ws_resumen[f'A{fila}'] = 'Recaudación Mínima'
        ws_resumen[f'B{fila}'] = self.resultados['recaudacion_min']
        ws_resumen[f'B{fila}'].number_format = '$#,##0.00'
        
        fila += 1
        ws_resumen[f'A{fila}'] = 'Recaudación Máxima'
        ws_resumen[f'B{fila}'] = self.resultados['recaudacion_max']
        ws_resumen[f'B{fila}'].number_format = '$#,##0.00'
        
        fila += 1
        ws_resumen[f'A{fila}'] = 'Ganancia Neta Promedio'
        ws_resumen[f'B{fila}'] = self.resultados['ganancia_promedio']
        ws_resumen[f'B{fila}'].number_format = '$#,##0.00'
        
        fila += 1
        ws_resumen[f'A{fila}'] = 'Sillas Necesarias (Máximo)'
        ws_resumen[f'B{fila}'] = self.resultados['max_sillas_necesarias']
        
        fila += 1
        ws_resumen[f'A{fila}'] = 'Sillas Promedio'
        ws_resumen[f'B{fila}'] = self.resultados['sillas_promedio']
        ws_resumen[f'B{fila}'].number_format = '0.00'
        
        fila += 1
        ws_resumen[f'A{fila}'] = 'Refrigerios Promedio por Día'
        ws_resumen[f'B{fila}'] = self.resultados['refrigerios_promedio']
        ws_resumen[f'B{fila}'].number_format = '0.00'
        
        fila += 1
        ws_resumen[f'A{fila}'] = 'Días con 5+ Refrigerios'
        ws_resumen[f'B{fila}'] = self.resultados['dias_5_o_mas_refrigerios']
        
        fila += 1
        ws_resumen[f'A{fila}'] = 'Probabilidad de 5+ Refrigerios'
        ws_resumen[f'B{fila}'] = self.resultados['prob_5_o_mas_refrigerios']
        ws_resumen[f'B{fila}'].number_format = '0.00%'
        
        # Ajustar ancho de columnas
        ws_resumen.column_dimensions['A'].width = 35
        ws_resumen.column_dimensions['B'].width = 20
        
        # Hoja 2: Resultados Diarios
        ws_diarios = wb.create_sheet("Resultados Diarios")
        
        # Título
        ws_diarios['A1'] = 'RESULTADOS DIARIOS'
        ws_diarios['A1'].font = titulo_font
        ws_diarios['A1'].fill = titulo_fill
        ws_diarios['A1'].alignment = center_alignment
        ws_diarios.merge_cells('A1:G1')
        
        # Encabezados
        headers = ['Día', 'Recaudación', 'Costo Refrig.', 'Ganancia Neta', 
                   'Clientes Atend.', 'Refrigerios', 'Sillas Usadas']
        for col, header in enumerate(headers, start=1):
            cell = ws_diarios.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Datos
        for idx, resultado in enumerate(self.resultados['resultados_diarios'], start=4):
            ws_diarios.cell(row=idx, column=1).value = resultado['dia']
            ws_diarios.cell(row=idx, column=2).value = resultado['recaudacion']
            ws_diarios.cell(row=idx, column=2).number_format = '$#,##0.00'
            ws_diarios.cell(row=idx, column=3).value = resultado['costo_refrigerios']
            ws_diarios.cell(row=idx, column=3).number_format = '$#,##0.00'
            ws_diarios.cell(row=idx, column=4).value = resultado['ganancia_neta']
            ws_diarios.cell(row=idx, column=4).number_format = '$#,##0.00'
            ws_diarios.cell(row=idx, column=5).value = resultado['clientes_atendidos']
            ws_diarios.cell(row=idx, column=6).value = resultado['clientes_con_refrigerio']
            ws_diarios.cell(row=idx, column=7).value = resultado['max_sillas_necesarias']
            
            # Aplicar bordes
            for col in range(1, 8):
                ws_diarios.cell(row=idx, column=col).border = border
        
        # Ajustar ancho de columnas
        for col in range(1, 8):
            ws_diarios.column_dimensions[chr(64 + col)].width = 15
        
        # Hoja 3: Vector de Estado
        ws_vector = wb.create_sheet("Vector de Estado")
        
        # Título
        ws_vector['A1'] = 'VECTOR DE ESTADO - ÚLTIMO DÍA SIMULADO'
        ws_vector['A1'].font = titulo_font
        ws_vector['A1'].fill = titulo_fill
        ws_vector['A1'].alignment = center_alignment
        ws_vector.merge_cells('A1:X1')
        
        # Encabezados
        headers_vector = [
            "Iter", "Reloj (min)", "Evento", 
            "RND Llegada", "RND Asig", "RND Servicio",
            "Prox. Llegada", "Prox. Fin Apr", "Prox. Fin VetA", "Prox. Fin VetB",
            "Estado Aprendiz", "Cliente Aprendiz", "Cola Aprendiz",
            "Estado Vet A", "Cliente Vet A", "Cola Vet A",
            "Estado Vet B", "Cliente Vet B", "Cola Vet B",
            "Clientes Atend.", "Recaud. Acum", "Costo Refrig", "Refrig. Entregados", "Max Cola"
        ]
        
        for col, header in enumerate(headers_vector, start=1):
            cell = ws_vector.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Obtener vector de estado filtrado
        hora_inicio = self.spin_hora_inicio.value()
        num_filas = self.spin_num_filas.value()
        filas_vector = self.ultima_simulacion.obtener_vector_estado_filtrado(hora_inicio, num_filas)
        
        # Datos del vector
        for idx, fila in enumerate(filas_vector, start=4):
            ws_vector.cell(row=idx, column=1).value = fila.iteracion
            ws_vector.cell(row=idx, column=2).value = fila.reloj
            ws_vector.cell(row=idx, column=2).number_format = '0.00'
            ws_vector.cell(row=idx, column=3).value = fila.evento
            
            # RNDs
            rnd_llegada = fila.rnd_evento.get('llegada', '')
            ws_vector.cell(row=idx, column=4).value = rnd_llegada if rnd_llegada != '' else '-'
            if rnd_llegada != '':
                ws_vector.cell(row=idx, column=4).number_format = '0.0000'
            
            rnd_asig = fila.rnd_evento.get('asignacion_peluquero', '')
            ws_vector.cell(row=idx, column=5).value = rnd_asig if rnd_asig != '' else '-'
            if rnd_asig != '':
                ws_vector.cell(row=idx, column=5).number_format = '0.0000'
            
            rnd_serv = fila.rnd_evento.get('tiempo_servicio', '')
            ws_vector.cell(row=idx, column=6).value = rnd_serv if rnd_serv != '' else '-'
            if rnd_serv != '':
                ws_vector.cell(row=idx, column=6).number_format = '0.0000'
            
            # Próximos eventos
            ws_vector.cell(row=idx, column=7).value = fila.proximo_llegada if fila.proximo_llegada > 0 else '-'
            if fila.proximo_llegada > 0:
                ws_vector.cell(row=idx, column=7).number_format = '0.00'
            
            ws_vector.cell(row=idx, column=8).value = fila.proximo_fin_aprendiz if fila.proximo_fin_aprendiz > 0 else '-'
            if fila.proximo_fin_aprendiz > 0:
                ws_vector.cell(row=idx, column=8).number_format = '0.00'
            
            ws_vector.cell(row=idx, column=9).value = fila.proximo_fin_veterano_a if fila.proximo_fin_veterano_a > 0 else '-'
            if fila.proximo_fin_veterano_a > 0:
                ws_vector.cell(row=idx, column=9).number_format = '0.00'
            
            ws_vector.cell(row=idx, column=10).value = fila.proximo_fin_veterano_b if fila.proximo_fin_veterano_b > 0 else '-'
            if fila.proximo_fin_veterano_b > 0:
                ws_vector.cell(row=idx, column=10).number_format = '0.00'
            
            # Estados de peluqueros
            ws_vector.cell(row=idx, column=11).value = fila.estado_aprendiz
            ws_vector.cell(row=idx, column=12).value = fila.cliente_aprendiz
            ws_vector.cell(row=idx, column=13).value = fila.cola_aprendiz
            
            ws_vector.cell(row=idx, column=14).value = fila.estado_veterano_a
            ws_vector.cell(row=idx, column=15).value = fila.cliente_veterano_a
            ws_vector.cell(row=idx, column=16).value = fila.cola_veterano_a
            
            ws_vector.cell(row=idx, column=17).value = fila.estado_veterano_b
            ws_vector.cell(row=idx, column=18).value = fila.cliente_veterano_b
            ws_vector.cell(row=idx, column=19).value = fila.cola_veterano_b
            
            # Acumuladores
            ws_vector.cell(row=idx, column=20).value = fila.clientes_atendidos
            ws_vector.cell(row=idx, column=21).value = fila.recaudacion_acum
            ws_vector.cell(row=idx, column=21).number_format = '$#,##0.00'
            ws_vector.cell(row=idx, column=22).value = fila.costo_refrigerios_acum
            ws_vector.cell(row=idx, column=22).number_format = '$#,##0.00'
            ws_vector.cell(row=idx, column=23).value = fila.clientes_con_refrigerio
            ws_vector.cell(row=idx, column=24).value = fila.max_cola_total
            
            # Aplicar bordes
            for col in range(1, 25):
                ws_vector.cell(row=idx, column=col).border = border
                ws_vector.cell(row=idx, column=col).alignment = Alignment(horizontal='center')
        
        # Ajustar ancho de columnas
        for col in range(1, 25):
            ws_vector.column_dimensions[chr(64 + col) if col <= 26 else f'A{chr(64 + col - 26)}'].width = 12
        
        # Guardar archivo
        wb.save(archivo)


def main():
    """Función principal"""
    app = QApplication(sys.argv)
    ventana = PeluqueriaVIPApp()
    ventana.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
